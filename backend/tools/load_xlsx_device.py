from openpyxl import load_workbook
from tabulate import tabulate


class Sheet:
    def __init__(self, filename):
        self.filename = filename
        self.dados = []
        self.number_lines = 0
        self.number_columns = 0
        self.cabecalhos = None
        self.ws = self.load_sheet()

    def get_number_columns(self):
        self.number_columns = len(self.cabecalhos)

    def get_name_columns(self):
        self.load_cabecalhos()
        for name in self.cabecalhos:
            print(name)
        print(self.cabecalhos)

    def load_sheet(self):
        wb = load_workbook(filename)
        return wb.active  # ou wb["NomeDaPlanilha"]

    def load_cabecalhos(self):
        # Pega os valores da primeira linha (cabeçalhos)
        self.cabecalhos = [celula.value for celula in self.ws[1]]

    def read_filename(self):
        for line in self.ws.iter_rows(min_row=2, values_only=True):
            self.number_lines += 1
            self.dados.append(line)

    def report(self):
        self.read_filename()
        self.load_cabecalhos()
        # adiciona os cabeçalhos como última linha também
        dados_com_rodape = self.dados + [self.cabecalhos]
        print(tabulate(dados_com_rodape, headers=self.cabecalhos, tablefmt="grid"))
        print(f'Planilha com {self.number_lines} linhas e {len(self.cabecalhos)} colunas!')

    def report_with_columns(self,desired_columns):
        ...


"""
ACTIVE - SUSPENSED - CANCELED
04/7 - Zé(Ative)lat=900 log=900 - Pedro(SUSPENSED)
05/7 - Zé(Canceled) - Pedro(Active)
not Canceled
update if SUSPENSED -> Active

"""


if '__main__' == __name__:
    filename = '../files/device.xlsx'
    desired_columns = [
        'CdrType',
        ' ThingName ',
        ' ThingsGroupName ',
        ' sessionCreationTime ',
        ' IMSI ',
        ' Usage ',
    ]
    """
    TABELA FILENAME DEVICE
    DeviceName só nrs
    DevicesProfileName
    MobileSegments (
    Default Mobile Segments for APN: internet   Default Mobile Segments for APN: apnname1
    internet - apnname1
    )
    MobileIdentities(IMSI)
    IMEI
    MSISDN
    MnoName
    CellId
    Lac
    State   
    """
    sheet = Sheet(filename)
    sheet.load_cabecalhos()
    # sheet.get_name_columns()
    sheet.report()
    # sheet.report_with_columns(desired_columns)


    # Limpa os espaços de ambas as listas para comparar certinho
    # colunas_desejadas_limpas = [c.strip() for c in colunas_desejadas]
    # # cabecalhos_limpos = [c.strip() if c else c for c in cabecalhos]
    # cabecalhos_limpos = [c.strip() if c else c for c in sheet.cabecalhos]
    #
    # # Pega os índices das colunas desejadas, na ordem em que aparecem na planilha
    # indices = [i for i, c in enumerate(cabecalhos_limpos) if c in colunas_desejadas_limpas]
    #
    # # Cabeçalhos finais (só os que você quer, na ordem da planilha)
    # # cabecalhos_filtrados = [cabecalhos[i] for i in indices]
    # cabecalhos_filtrados = [sheet.cabecalhos[i] for i in indices]
    #
    # # Filtra os dados de cada linha, pegando só as colunas desejadas
    # dados_filtrados = []
    # # for linha in ws.iter_rows(min_row=2, values_only=True):
    # for linha in sheet.ws.iter_rows(min_row=2, values_only=True):
    #     dados_filtrados.append([linha[i] for i in indices])
    #
    # print(tabulate(dados_filtrados, headers=cabecalhos_filtrados, tablefmt="grid"))


    # sheet.get_name_columns()
    # sheet.report()
    # r = "Id - Timestamp - SessionId - IMSI - IMEI - ThingId - RealUsage - Usage - UOM".split('-')
    # print(r)

# "Id - Timestamp - SessionId - IMSI - IMEI - ThingId - RealUsage - Usage - UOM"