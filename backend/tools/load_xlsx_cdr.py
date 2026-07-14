import os
import django

# 1. Altere 'nome_do_seu_projeto' para o nome da pasta principal do seu projeto Django
os.environ.setdefault(
    'DJANGO_SETTINGS_MODULE',
    'kernel.settings'
)

# 2. Inicializa o Django
django.setup()

from openpyxl import load_workbook
from tabulate import tabulate

from cdr.models import Organization


class Sheet:
    def __init__(self, filename):
        self.filename = filename
        self.dados = []
        self.number_lines = 0
        self.number_columns = 0
        self.cabecalhos = None
        self.ws = self.load_sheet()

    def get_number_columns(self):
        self.number_columns = len(self.cabecalhos)

    def get_name_columns(self):
        self.load_cabecalhos()
        for name in self.cabecalhos:
            print(name)
        print(self.cabecalhos)

    def load_sheet(self):
        wb = load_workbook(self.filename)
        return wb.active  # ou wb["NomeDaPlanilha"]

    def load_cabecalhos(self):
        # Pega os valores da primeira linha (cabeçalhos)
        self.cabecalhos = [celula.value for celula in self.ws[1]]

    def read_filename(self):
        if self.dados:
            return
        for line in self.ws.iter_rows(min_row=2, values_only=True):
            self.number_lines += 1
            self.dados.append(line)

    def report(self):
        self.read_filename()
        self.load_cabecalhos()
        # adiciona os cabeçalhos como última linha também
        dados_com_rodape = self.dados + [self.cabecalhos]
        print(tabulate(dados_com_rodape, headers=self.cabecalhos, tablefmt="grid"))
        print(f'Planilha com {self.number_lines} linhas e {len(self.cabecalhos)} colunas!')

    def report_with_columns(self, desired_columns):
        self.read_filename()
        # adiciona os cabeçalhos como última linha também
        colunas_desejadas = [c.strip() for c in desired_columns]
        cabecalhos_limpos = [c.strip() for c in self.cabecalhos]
        indices = [i for i, c in enumerate(cabecalhos_limpos) if c in colunas_desejadas]

        """
        BD = CSV
        ---------
        orgId = OrganizationId
        orgname = OrganizationName
        mnoid = MNOId
        mnoname = MnoName
        customerid = CustomerId
        customername = CustomerName
        networkproviderid = NetworkProviderId
        networkprovidername = NetworkProviderName
        priceplanid = PricePlanId
        priceplanname = PricePlanName
        thinggroupid = ThingsGroupId depois do "ThingsGroupId_"
        thinggroupname = ThingsGroupName
        thingnameraw = ThingsName só nros
        thingname = ThingsName só nros
        thinggroupidraw = ThingsGroupId
        """

        cabecalhos_filtrados = [self.cabecalhos[i] for i in indices]
        dados_com_rodape = self.dados + [self.cabecalhos]
        dados_filtrados = [[linha[i] for i in indices] for linha in dados_com_rodape]
        print(tabulate(dados_filtrados, headers=cabecalhos_filtrados, tablefmt="grid"))
        print(f'Planilha com {self.number_lines} linhas e {len(indices)} colunas selecionadas!')

    def prepare_for_db(self, column_mapping):
        self.read_filename()
        self.load_cabecalhos()

        cabecalhos_limpos = [c.strip() for c in self.cabecalhos]
        colunas_excel = list(column_mapping.values())
        indices = {col: i for i, col in enumerate(cabecalhos_limpos) if col in colunas_excel}

        mnos, customers, network_providers, price_plans, things = {}, {}, {}, {}, {}
        vistos = set()
        organizacoes_unicas = []

        def get(linha, col):
            if col not in indices:
                return ''
            val = linha[indices[col]]
            return str(val).strip() if val else ''

        for linha in self.dados:
            orgid = get(linha, 'OrganizationId')
            orgname = get(linha, 'OrganizationName')

            # Se o org_id ainda não foi processado, adicionamos na lista única
            if orgid not in vistos:
                vistos.add(orgid)
                Organization.objects.update_or_create(
                    orgid=orgid,  # Critério para buscar se já existe
                    defaults={
                        'orgname': orgname
                    }  # O que atualizar/definir caso crie
                )

            mnoid = get(linha, 'MNOId')
            mnoname = get(linha, 'MnoName')
            if mnoid and mnoid not in mnos:
                mnos[mnoid] = {
                    'mnoid': mnoid,
                    'mnoname': mnoname
                }

            np_id = get(linha, 'NetworkProviderId')
            np_name = get(linha, 'NetworkProviderName')
            if np_id and np_id not in network_providers:
                network_providers[np_id] = {
                    'netproviderid': np_id,
                    'netprovidername': np_name
                }

            pp_id = get(linha, 'PricePlanId')
            pp_name = get(linha, 'PricePlanName')
            if pp_id and pp_id not in price_plans:
                price_plans[pp_id] = {
                    'priceplanid': pp_id,
                    'priceplanname': pp_name
                }

            tg_id_raw = get(linha, 'ThingsGroupId')
            tg_id = tg_id_raw.split('_', 1)[1] if '_' in tg_id_raw else tg_id_raw
            tg_name = get(linha, 'ThingsGroupName')
            tg_name = ''.join(
                    c for c in tg_name if c.isdigit()
                )
            if tg_id and tg_id not in things:
                things[tg_id] = {
                    'thingsgroupid': tg_id,
                    'thingsgroupname': tg_name,
                }

            cust_id = get(linha, 'CustomerId')
            cust_name = get(linha, 'CustomerName')
            if cust_id and cust_id not in customers:
                customers[cust_id] = {
                    'customerid': cust_id,
                    'customername': cust_name,
                    '_networkprovider_id': np_id,
                    '_priceplan_id': pp_id,
                    '_thing_id': tg_id,
                }

        return items
        # return {
        #     'mno': list(mnos.values()),
        #     'networkprovider': list(network_providers.values()),
        #     'priceplan': list(price_plans.values()),
        #     'thing': list(things.values()),
        #     'customer': list(customers.values()),
            # 'organization': organizations,
        # }


"""
ACTIVE - SUSPENSED - CANCELED
04/7 - Zé(Ative)lat=900 log=900 - Pedro(SUSPENSED)
05/7 - Zé(Canceled) - Pedro(Active)
not Canceled
update if SUSPENSED -> Active

"""

if '__main__' == __name__:
    # filename = '../files/cdr.xlsx'
    filename = os.path.join(os.path.dirname(__file__), '..', 'files', 'cdr.xlsx')
    sheet = Sheet(filename)
    sheet.load_cabecalhos()
    # sheet.get_name_columns()
    # sheet.report()
    """
    -----
    model Organization
    orgid = OrganizationId
    orgname = OrganizationName
    ----
    model Mno
    mnoid = MNOId
    mnoname = MnoName
    ----
    model Customer
    customerid = CustomerId
    customername = CustomerName
    ----
    model NetworkProvider
    netproviderid = NetworkProviderId
    netprovidername = NetworkProviderName
    ----
    model PricePlan
    priceplanid = PricePlanId
    priceplanname = PricePlanName
    ----
    model Thing
    thinggroupid = ThingsGroupId
    thinggroupname = ThingsGroupName
    thingnameraw = 
    thinggroupidraw = 
    
    """
    column_mapping = {
        # Organization
        'orgid': 'OrganizationId',
        'orgname': 'OrganizationName',
        # MNO
        'mnoid': 'MNOId',
        'mnoname': 'MnoName',
        # Customer
        'customerid': 'CustomerId',
        'customername': 'CustomerName',
        # Network Provide
        'networkproviderid': 'NetworkProviderId',
        'networkprovidername': 'NetworkProviderName',
        # PricePlan
        'priceplanid': 'PricePlanId',
        'priceplanname': 'PricePlanName',
        # Thing
        'thinggroupid': 'ThingsGroupId',
        'thinggroupname': 'ThingsGroupName',
    }
    desired_columns = list(column_mapping.values())
    sheet.report_with_columns(desired_columns)
    dados_para_db = sheet.prepare_for_db(column_mapping)
    print(dados_para_db)
    """
    TABELA FILENAME DEVICE
    DeviceName só nrs
    DevicesProfileName
    MobileSegments (
    Default Mobile Segments for APN: internet   Default Mobile Segments for APN: apnname1
    internet - apnname1
    )
    MobileIdentities(IMSI)
    IMEI
    MSISDN
    MnoName
    CellId
    Lac
    State   
    """

    """
    -----------------------------------------------------------------------
    'OrganizationId', 
    'OrganizationName', 
    'ThingId', 
    'ThingName', 
    'ThingsGroupId', 
    'ThingsGroupName', 
    'CustomerId', 
    'CustomerName', 
    'PricePlanId', 
    'PricePlanName', 
    'NetworkProviderId', 
    'NetworkProviderName', 
    'MNOId', 
    'MnoName', 
    'sessionCreationTime', 
    'IMSI', 
    'MSISDN', 
    'IMEI', 
    'TADIG', 
    'Usage', 
    'UOM'
    """
